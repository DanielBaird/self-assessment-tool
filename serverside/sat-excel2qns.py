#!/usr/bin/env python

import argparse
import sys

try:
	import configparser # python 3.x name
except ImportError:
	import ConfigParser as configparser # puython 2.x name

import openpyxl



# -------------------------------------------------------------------
def read_config(config_file_path):
	cp = configparser.SafeConfigParser()
	cp.read(config_file_path)
	return cp

# -------------------------------------------------------------------
def open_excel_sheet(excel_file_path, sheet_name=None):
	wb = openpyxl.load_workbook(filename=excel_file_path, read_only=True)
	sheet_count = len(wb.sheetnames)

	if sheet_count < 1:
		sys.exit('no sheets in excel file provided.')

	if sheet_name == None:
		sheet_name = wb.sheetnames[0]

	if sheet_name in wb.sheetnames:
		return wb[sheet_name]
	else:
		sys.exit('Worksheet named "' + sheet_name + '" not found in Workbook.')

# -------------------------------------------------------------------
def process_sheet(sheet, cfg):
	questions = {}

	# shortcuts to col indexes
	cols = {}
	headerrow = cfg.getint('layout', 'headerrow')
	headers = sheet.iter_rows(min_row=headerrow, max_row=headerrow)
	# there's just one row, we're for-ing it coz its in a generator
	for r in headers:
		for c in r: # for each cell in this header row
			for col in map(lambda item: item[0], cfg.items('columnNames')):
				if c.value == cfg.get('columnNames', col):
					cols[col] = c.column

	for col in cols:
		print('At column ' + str(cols[col]) + ', found ' + str(col) + ' ("' + cfg.get('columnNames', col) + '")')

# -------------------------------------------------------------------
parser = argparse.ArgumentParser(
	description='Converts questions in an Excel file into a question data file suitable for SAT display.',
	# usage='%(prog)s [options] <input excelfile> <output questionfile>'
)
parser.add_argument('excelfile', help='Excel document input containing questions')
parser.add_argument('questionfile', help='output question data file (default: qns.json)', nargs='?', default='qns.json')
parser.add_argument('--config', help='specify config file (default: ./satconfig.cfg)', default='satconfig.cfg')
parser.add_argument('--sheet', help='specify the worksheet name (default: the first sheet in the workbook)')

args = parser.parse_args()
cfg = read_config(args.config)
sheet = open_excel_sheet(args.excelfile, args.sheet)

process_sheet(sheet, cfg)

